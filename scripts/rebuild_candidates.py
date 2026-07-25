"""Rebuild candidates.json from application Excel + enrich insights."""
from __future__ import annotations

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "candidates.json"

XLSX_CANDIDATES = [
    Path(r"C:\Users\prana\OneDrive\Desktop\SPORTSCOM\_                Application Round (1-113) (1).xlsx"),
    Path(
        r"C:\Users\prana\AppData\Local\Packages\5319275A.WhatsAppDesktop_cv1g1gvanyjgm\LocalState\sessions\3DFE4DA32114A2B34ED40B480FCB5E2249621FE5\transfers\2026-30\_                Application Round (1-113).xlsx"
    ),
]

SPORTS = [
    "Cricket",
    "Table Tennis",
    "Swimming",
    "Badminton",
    "Volleyball",
    "Basketball",
    "Throw ball",
    "Athletics",
    "Football",
    "Lawn Tennis",
    "Squash",
]

LEVEL_PATTERNS = [
    ("International", r"\b(international|world\s*cup|olympics?|asia(?:n)?\s*(?:level|games|championship)?)\b"),
    ("National", r"\b(national|nationals|all[\s-]?india|cbse\s*nationals?)\b"),
    ("State", r"\b(state(?:[\s-]?level)?|state\s*(?:championship|tournament|team|games)|cbse\s*clusters?)\b"),
    ("Zonal / Regional", r"\b(zonal|zone|regional|inter[\s-]?zone)\b"),
    ("District", r"\b(district|inter[\s-]?district)\b"),
    ("University", r"\b(university|inter[\s-]?university|varsity)\b"),
    ("College", r"\b(college|inter[\s-]?college|undergrad)\b"),
    ("School", r"\b(school|inter[\s-]?school|cbse|clusters?)\b"),
]

TRAIT_RULES = [
    ("Event organising", r"\b(organiz|organis|event|coordina|volunteer|manage(d|ment)?|logistics|conduct)\b"),
    ("Leadership", r"\b(captain|lead(er|ership)?|headed|president|vice[\s-]?captain)\b"),
    ("Team spirit", r"\b(team(mate|work| player)?|together|squad|collective)\b"),
    ("Participation & culture", r"\b(participat|encourag|culture|community|involve|engage|inclus)\b"),
    ("Competitive drive", r"\b(compet(e|itive|ition)|win(ning)?|championship|tournament|medal|trophy|final)\b"),
    ("Design / content", r"\b(design|canva|edit(ing|or)?|content|poster|video|capcut|premiere|instagram|reel)\b"),
    ("Playing ability", r"\b(play(ed|ing|er)?|represented|proficien|skill(ed)?)\b"),
]


def find_xlsx() -> Path:
    for p in XLSX_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("Application Round xlsx not found")


def to_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def pick_sports(row: dict):
    scores = {}
    for s in SPORTS:
        a = to_int(row.get(s))
        b = to_int(row.get(s + "2"))
        vals = [x for x in (a, b) if x is not None]
        if vals:
            scores[s] = max(vals)
    top = sorted([(k, v) for k, v in scores.items() if v >= 3], key=lambda x: (-x[1], x[0]))
    return [{"sport": k, "level": v} for k, v in top], scores


def split_interests(raw) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;,\n]+", str(raw))
    return [p.strip() for p in parts if p.strip()]


def pick_design(row: dict) -> dict:
    canva = to_int(row.get("Canva"))
    canva2 = to_int(row.get("Canva2"))
    video = to_int(row.get("Video Editing (Capcut/ Premier pro/ etc)"))
    video2 = to_int(row.get("Video Editing (Capcut/ Premier pro/ etc)2"))
    interests = split_interests(row.get("What type of work interests you the most")) or split_interests(
        row.get("What type of work interests you the most2")
    )
    canva_final = max([x for x in (canva, canva2) if x is not None], default=None)
    video_final = max([x for x in (video, video2) if x is not None], default=None)
    why_content = str(row.get("Why do you want to join the Content & Design team") or "").strip()
    why_event = str(row.get("Why do you want to join the Event Coordinator Team") or "").strip()
    contribute_both = str(row.get("How can you contribute yourself to both of these verticals?") or "").strip()
    return {
        "canva": canva_final,
        "videoEditing": video_final,
        "workInterests": interests,
        "whyContent": why_content,
        "whyEvent": why_event,
        "contributeBoth": contribute_both,
    }


def wants_design(vertical: str) -> bool:
    v = (vertical or "").lower()
    return "design" in v or "content" in v or "both" in v


def sentences(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", (text or "").strip())
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+|\n+|;\s*", text)
    return [p.strip(" -•\t") for p in parts if len(p.strip()) > 12]


def shorten(s: str, n: int = 140) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) <= n:
        return s
    cut = s[:n].rsplit(" ", 1)[0]
    return cut.rstrip(",;:") + "..."


def bulletize(text: str, max_bullets: int = 4) -> list[str]:
    sents = sentences(text)
    if not sents:
        t = (text or "").strip()
        return [shorten(t, 160)] if t else []
    scored = []
    for s in sents:
        if s and s[0].islower() and not s.lower().startswith(("i ", "i'", "we ", "my ")):
            continue
        score = 0
        low = s.lower()
        for w in [
            "represent",
            "won",
            "captain",
            "state",
            "national",
            "district",
            "college",
            "tournament",
            "medal",
            "organiz",
            "particip",
            "lead",
            "play",
            "design",
            "content",
            "canva",
            "video",
            "contribut",
            "would",
        ]:
            if w in low:
                score += 2
        score += min(len(s) // 40, 3)
        scored.append((score, s))
    if not scored:
        scored = [(1, s) for s in sents]
    scored.sort(key=lambda x: -x[0])
    picked: list[str] = []
    for _, s in scored:
        if len(picked) >= max_bullets:
            break
        if any(s[:40].lower() in p.lower() or p[:40].lower() in s.lower() for p in picked):
            continue
        picked.append(shorten(s, 150))
    return picked[:max_bullets]


def extract_levels(text: str) -> list[str]:
    low = text.lower()
    found: list[str] = []
    for label, pat in LEVEL_PATTERNS:
        if re.search(pat, low, re.I) and label not in found:
            found.append(label)
    return found


def level_mentions(text: str) -> list[str]:
    hits: list[str] = []
    for s in sentences(text):
        low = s.lower()
        if not re.search(
            r"\b(represented|national|state|district|university|college|school|cbse|clusters?|championship|tournament|medal|won|winner|captain)\b",
            low,
        ):
            continue
        snippet = shorten(s, 130)
        if snippet.lower() not in [h.lower() for h in hits]:
            hits.append(snippet)
    return hits[:5]


def strengths(c: dict) -> list[str]:
    items: list[str] = []
    tops = sorted(c.get("topSports") or [], key=lambda x: -x["level"])
    for s in tops[:4]:
        if s["level"] >= 4:
            items.append(f"Strong at {s['sport']} (self-rated {s['level']}/5)")
        elif s["level"] == 3:
            items.append(f"Solid {s['sport']} base (self-rated 3/5)")
    design = c.get("design") or {}
    if design.get("canva") and design["canva"] >= 3:
        items.append(f"Canva (self-rated {design['canva']}/5)")
    if design.get("videoEditing") and design["videoEditing"] >= 3:
        items.append(f"Video editing (self-rated {design['videoEditing']}/5)")
    for interest in (design.get("workInterests") or [])[:3]:
        items.append(f"Interested in: {interest}")
    blob = " ".join(
        [
            c.get("achievements") or "",
            c.get("cultureAnswer") or "",
            c.get("favoriteMemory") or "",
            design.get("whyContent") or "",
            design.get("whyEvent") or "",
            design.get("contributeBoth") or "",
        ]
    )
    for label, pat in TRAIT_RULES:
        if re.search(pat, blob, re.I):
            if label not in items and not any(label.lower() in i.lower() for i in items):
                items.append(label)
    if c.get("vertical"):
        items.append(f"Applied for: {c['vertical']}")
    out: list[str] = []
    for i in items:
        if i not in out:
            out.append(i)
    return out[:8]


def build_insight(c: dict) -> dict:
    design = c.get("design") or {}
    parts = [
        c.get("achievements") or "",
        c.get("otherSports") or "",
        c.get("cultureAnswer") or "",
        c.get("favoriteMemory") or "",
        design.get("whyContent") or "",
        design.get("whyEvent") or "",
        design.get("contributeBoth") or "",
    ]
    blob = ". ".join(p.strip().rstrip(".") for p in parts if p and str(p).strip())
    insight = {
        "goodAt": strengths(c),
        "competitionLevels": extract_levels(blob),
        "levelMentions": level_mentions(blob),
        "sportsMentioned": [],
        "achievementsBullets": bulletize(c.get("achievements") or "", 4)
        if (c.get("achievements") or "").strip()
        else [],
        "sportsComBullets": bulletize(c.get("cultureAnswer") or "", 4),
        "memoryBullets": bulletize(c.get("favoriteMemory") or "", 3),
        "designBullets": bulletize(design.get("whyContent") or design.get("contributeBoth") or "", 3),
        "summary": "",
    }
    bits = []
    if insight["goodAt"]:
        bits.append(insight["goodAt"][0])
    if insight["competitionLevels"]:
        bits.append("Mentioned " + ", ".join(insight["competitionLevels"][:3]) + " level")
    insight["summary"] = " · ".join(bits) if bits else "Form filled; limited detail shared."
    return insight


def main() -> None:
    import openpyxl

    src = find_xlsx()
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    candidates = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if not row.get("ID") and not row.get("Name2") and not row.get("Name"):
            continue
        name = (row.get("Name2") or row.get("Name") or "").strip()
        if not row.get("Name2") and name:
            name = re.sub(r"\s+\d{10,}$", "", name).strip()
        top_sports, all_scores = pick_sports(row)
        other = row.get("Any other Sport other than the ones mentioned above? (State your proficiency as well)") or row.get(
            "Any other Sport other than the ones mentioned above? (State your proficiency as well)2"
        )
        vertical = str(
            row.get("Which Vertical would you prefer? Select both if you are interested in both the verticals") or ""
        ).strip()
        design = pick_design(row)
        c = {
            "id": int(row["ID"]) if row.get("ID") is not None else r - 1,
            "name": name,
            "email": (row.get("TAPMI Email ID") or row.get("Email") or "").strip(),
            "vertical": vertical,
            "achievements": str(row.get("Any of your Major achievements in Sports?(Not Mandatory)") or "").strip(),
            "otherSports": str(other or "").strip(),
            "cultureAnswer": str(
                row.get(
                    "What does being part of the Sports Committee mean to you, and how would you contribute to building a strong sports culture at TAPMI?(Max: 500 words)"
                )
                or ""
            ).strip(),
            "favoriteMemory": str(
                row.get("Your favorite sports memory?(Watching or Playing)(Max:500 words)") or ""
            ).strip(),
            "topSports": top_sports,
            "allScores": all_scores,
            "design": design,
            "wantsDesign": wants_design(vertical),
        }
        c["insight"] = build_insight(c)
        candidates.append(c)

    OUT.write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    design_n = sum(1 for c in candidates if c["wantsDesign"])
    print(f"exported {len(candidates)} candidates ({design_n} with design/content interest)")
    sample = next(c for c in candidates if c["wantsDesign"])
    print(sample["name"], sample["vertical"], sample["design"])


if __name__ == "__main__":
    main()
